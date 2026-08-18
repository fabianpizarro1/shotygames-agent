#!/usr/bin/env python3
"""Genera el libro de costos de CandyShots (INSUMOS -> RECETAS -> MENU)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()

NARANJA = "F05A28"
CREMA = "FFF4EE"
GRIS = "F7F7F7"
H_FONT = Font(bold=True, color="FFFFFF", size=11)
H_FILL = PatternFill("solid", fgColor=NARANJA)
SEC_FILL = PatternFill("solid", fgColor=CREMA)
thin = Side(style="thin", color="DDDDDD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '"$"#,##0.0000'
MONEY2 = '"$"#,##0.00'
PCT = '0.0%'


def header(ws, row, cols, widths):
    for i, (c, w) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30


# ───────────────────────────── INSTRUCCIONES ─────────────────────────────
ws = wb.active
ws.title = "INSTRUCCIONES"
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 110
ws["B2"] = "CANDYSHOTS — CALCULADORA DE COSTOS"
ws["B2"].font = Font(bold=True, size=18, color=NARANJA)
ws["B3"] = "Agosto 2026 · Todo se recalcula solo. Tú solo tocas la columna de precios."
ws["B3"].font = Font(italic=True, size=11, color="666666")

pasos = [
    ("", ""),
    ("CÓMO SE USA", "header"),
    ("1. Andá a la hoja INSUMOS. Es la lista de todo lo que se compra.", ""),
    ("2. En la columna 'Precio de compra' pone lo que TE COSTÓ de verdad, con la factura en la mano.", ""),
    ("3. En 'Cantidad que rinde' pone cuánto trae esa compra (gramos, ml, unidades).", ""),
    ("   Ejemplo: comprás una libra de carne a $2.00 → Precio 2.00 · Cantidad 453.6 · Unidad g", ""),
    ("4. Listo. Las hojas RECETAS y MENÚ se recalculan solas.", ""),
    ("", ""),
    ("LAS 3 HOJAS", "header"),
    ("INSUMOS  →  Lista de compras. Acá ponés los precios reales. ES LA ÚNICA HOJA QUE TENÉS QUE LLENAR.", ""),
    ("RECETAS  →  Qué lleva cada plato y cuánto. Se puede agregar o quitar ingredientes libremente.", ""),
    ("MENÚ     →  Precio de venta, costo, food cost % y margen. Semáforo automático.", ""),
    ("", ""),
    ("EL SEMÁFORO", "header"),
    ("✅ SANO      food cost hasta 30%   — así debe estar todo", ""),
    ("🟡 LÍMITE    30% a 35%             — aceptable, vigilar", ""),
    ("🔴 REVISAR   más de 35%            — perdés plata o casi", ""),
    ("", ""),
    ("PARA AGREGAR UN PRODUCTO NUEVO", "header"),
    ("1. En RECETAS agregá una fila por cada ingrediente, escribiendo SIEMPRE el mismo nombre de producto.", ""),
    ("2. En MENÚ agregá una fila con ese mismo nombre exacto y su precio de venta.", ""),
    ("3. El costo se suma solo. Si el nombre no coincide exacto, el costo sale en cero.", ""),
    ("", ""),
    ("OJO CON ESTO", "header"),
    ("Las filas 'AJUSTE: extras sin costear' existen para que los totales cuadren con el PDF de agosto.", ""),
    ("Representan apanado, aceite, salsas y empaque que todavía no están costeados uno por uno.", ""),
    ("A medida que consigas esos precios reales, agregá el insumo y bajá o borrá la fila de AJUSTE.", ""),
    ("", ""),
    ("PENDIENTE: BEBIDAS", "header"),
    ("No hay ninguna bebida cargada porque no estaban en el menú que se pasó. Cuando las tengas se agregan.", ""),
]
r = 5
for texto, tipo in pasos:
    c = ws.cell(row=r, column=2, value=texto)
    if tipo == "header":
        c.font = Font(bold=True, size=12, color=NARANJA)
        c.fill = SEC_FILL
    else:
        c.font = Font(size=11)
    r += 1

# ───────────────────────────── INSUMOS ─────────────────────────────
ins = wb.create_sheet("INSUMOS")
ins["A1"] = "INSUMOS — poné acá los precios reales de compra"
ins["A1"].font = Font(bold=True, size=14, color=NARANJA)
ins["A2"] = "Solo tocás las columnas B y C. El costo unitario se calcula solo."
ins["A2"].font = Font(italic=True, size=10, color="666666")
header(ins, 4, ["INSUMO", "Precio de compra", "Cantidad que rinde", "Unidad",
                "COSTO UNITARIO", "Estado", "Nota"],
       [38, 16, 18, 10, 16, 18, 46])

INSUMOS = [
    ("— COMIDA —", None, None, None, "", ""),
    ("Papa", 3.40, 1000, "g", "✅ Confirmado", ""),
    ("Pollo crudo", 5.53, 1000, "g", "✅ Confirmado", "$0.83 por 150 g"),
    ("Carne molida", 2.00, 453.6, "g", "✅ Confirmado", "$2.00 la libra"),
    ("Pan de hamburguesa", 2.85, 6, "unidad", "✅ Confirmado", "Buscar uno más barato para la Clásica"),
    ("Cheddar laminado", 6.15, 1000, "g", "✅ Confirmado", "Lámina de 20 g = $0.12"),
    ("Huevo", 1.00, 7, "unidad", "✅ Confirmado", ""),
    ("Aro de cebolla congelado", 6.62, 1000, "g", "✅ Confirmado", "Aro de 25 g = $0.17"),
    ("Salchicha", 0.35, 1, "unidad", "✅ Confirmado", ""),
    ("Tocino", 0.25, 25, "g", "✅ Confirmado", "1 tira ≈ 25 g"),
    ("Maduro", 0.25, 1, "unidad", "✅ Confirmado", ""),
    ("Guacamole", 0.25, 1, "porción", "✅ Confirmado", "2 cucharadas"),
    ("Patacón", 0.10, 1, "unidad", "✅ Confirmado", ""),
    ("Chorizo", 6.90, 1000, "g", "⚠️ CONFIRMAR", "Estimado — el PDF lo marca como pendiente"),
    ("Cebolla caramelizada", 0.16, 1, "porción", "⚠️ CONFIRMAR", "Estimado"),
    ("Lechuga", 0, 1, "porción", "❌ FALTA", "Conseguir precio"),
    ("Tomate", 0, 1, "porción", "❌ FALTA", "Conseguir precio"),
    ("Salsas (porción)", 0, 1, "porción", "❌ FALTA", "Conseguir precio"),
    ("Apanado", 0, 1, "porción", "❌ FALTA", "Conseguir precio"),
    ("Aceite (por fritura)", 0, 1, "porción", "❌ FALTA", "Conseguir precio"),
    ("— EMPAQUE —", None, None, None, "", ""),
    ("Bandeja kraft", 0.04, 1, "unidad", "✅ Confirmado", "Consumo en el local"),
    ("Lonchera biodegradable 8\"", 0.23, 1, "unidad", "✅ Confirmado", "Delivery"),
    ("Cono de papel", 0, 1, "unidad", "❌ FALTA", "Conseguir precio"),
    ("Vaso 12oz + tapa + pitillo", 0.10, 1, "set", "⚠️ CONFIRMAR", "Estimado de mayo"),
    ("— GRANIZADOS —", None, None, None, "", ""),
    ("Jarabe concentrado", 10.00, 3785, "ml", "⚠️ CONFIRMAR", "1 galón · dilución 1:3 · 62.5 ml por vaso"),
    ("Hielo", 0.75, 2268, "g", "⚠️ CONFIRMAR", "Bolsa de 5 lb"),
    ("Leche entera", 1.00, 1000, "ml", "⚠️ CONFIRMAR", "Estimado de mayo"),
    ("Leche condensada", 4.50, 1000, "ml", "⚠️ CONFIRMAR", "Estimado de mayo"),
    ("Crema de leche", 3.00, 1000, "ml", "⚠️ CONFIRMAR", "Estimado de mayo"),
    ("Galleta Oreo", 3.20, 36, "unidad", "⚠️ CONFIRMAR", "Paquete de 36"),
    ("Nutella", 16.00, 1000, "g", "⚠️ CONFIRMAR", "Estimado de mayo"),
    ("Café soluble", 20.00, 1000, "g", "⚠️ CONFIRMAR", "Estimado de mayo"),
    ("Alcohol (aguardiente/vodka)", 7.00, 750, "ml", "⚠️ CONFIRMAR", "Porción de 30 ml"),
    ("— BEBIDAS —", None, None, None, "", ""),
    ("Cola", 0, 1, "unidad", "❌ FALTA", "No estaba en el menú entregado"),
    ("— AJUSTE TEMPORAL —", None, None, None, "", ""),
    ("AJUSTE: extras sin costear", 1.00, 1, "$", "❌ REEMPLAZAR",
     "Apanado + aceite + salsas + empaque. La cantidad en RECETAS es el monto en dólares."),
]

r = 5
for nombre, precio, cant, unidad, estado, nota in INSUMOS:
    if precio is None:
        c = ins.cell(row=r, column=1, value=nombre)
        c.font = Font(bold=True, size=11, color=NARANJA)
        for col in range(1, 8):
            ins.cell(row=r, column=col).fill = SEC_FILL
            ins.cell(row=r, column=col).border = BORDER
        r += 1
        continue
    ins.cell(row=r, column=1, value=nombre).font = Font(size=11)
    ins.cell(row=r, column=2, value=precio).number_format = MONEY2
    ins.cell(row=r, column=3, value=cant)
    ins.cell(row=r, column=4, value=unidad)
    f = ins.cell(row=r, column=5, value=f"=IFERROR(B{r}/C{r},0)")
    f.number_format = MONEY
    f.font = Font(bold=True)
    f.fill = PatternFill("solid", fgColor=GRIS)
    ins.cell(row=r, column=6, value=estado)
    ins.cell(row=r, column=7, value=nota).font = Font(size=9, color="777777")
    for col in range(1, 8):
        ins.cell(row=r, column=col).border = BORDER
    ins.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FFFDE7")
    ins.cell(row=r, column=3).fill = PatternFill("solid", fgColor="FFFDE7")
    r += 1

INS_LAST = r - 1
ins.freeze_panes = "A5"

for rng, color in [('"❌"', "FFCDD2"), ('"⚠️"', "FFF9C4"), ('"✅"', "C8E6C9")]:
    pass
ins.conditional_formatting.add(f"F5:F{INS_LAST}",
    CellIsRule(operator="containsText", formula=[f'NOT(ISERROR(SEARCH("FALTA",F5)))'],
               fill=PatternFill("solid", fgColor="FFCDD2")))

# ───────────────────────────── RECETAS ─────────────────────────────
rec = wb.create_sheet("RECETAS")
rec["A1"] = "RECETAS — qué lleva cada producto"
rec["A1"].font = Font(bold=True, size=14, color=NARANJA)
rec["A2"] = "Podés agregar o borrar filas. El nombre del producto debe coincidir EXACTO con la hoja MENÚ."
rec["A2"].font = Font(italic=True, size=10, color="666666")
header(rec, 4, ["PRODUCTO", "INSUMO", "Cantidad", "Unidad", "COSTO"], [34, 34, 12, 10, 14])

R = {
    "Bandeja Clásica": [("Papa", 150, "g"), ("Pollo crudo", 150, "g"), ("Salchicha", 1, "unidad"),
                        ("Bandeja kraft", 1, "unidad"), ("AJUSTE: extras sin costear", 0.44, "$")],
    "Bandeja Completa": [("Papa", 150, "g"), ("Pollo crudo", 150, "g"), ("Tocino", 25, "g"),
                         ("Maduro", 1, "unidad"), ("Bandeja kraft", 1, "unidad"),
                         ("AJUSTE: extras sin costear", 0.49, "$")],
    "Bandeja Suprema": [("Papa", 150, "g"), ("Pollo crudo", 150, "g"), ("Salchicha", 1, "unidad"),
                        ("Tocino", 25, "g"), ("Maduro", 1, "unidad"), ("Guacamole", 1, "porción"),
                        ("Patacón", 1, "unidad"), ("Bandeja kraft", 1, "unidad"),
                        ("AJUSTE: extras sin costear", 0.54, "$")],
    "Solitaria de Pollo": [("Papa", 180, "g"), ("Pollo crudo", 60, "g"), ("Cheddar laminado", 25, "g"),
                           ("AJUSTE: extras sin costear", 0.40, "$")],
    "Solitaria de Chorizo": [("Papa", 180, "g"), ("Chorizo", 60, "g"), ("Cheddar laminado", 25, "g"),
                             ("AJUSTE: extras sin costear", 0.40, "$")],
    "Pollo Bacon": [("Papa", 200, "g"), ("Pollo crudo", 60, "g"), ("Tocino", 25, "g"),
                    ("Cheddar laminado", 30, "g"), ("AJUSTE: extras sin costear", 0.41, "$")],
    "Chori Bacon": [("Papa", 200, "g"), ("Chorizo", 50, "g"), ("Tocino", 25, "g"),
                    ("Cheddar laminado", 30, "g"), ("AJUSTE: extras sin costear", 0.41, "$")],
    "Carne Bacon": [("Papa", 200, "g"), ("Carne molida", 60, "g"), ("Tocino", 25, "g"),
                    ("Cheddar laminado", 30, "g"), ("AJUSTE: extras sin costear", 0.41, "$")],
    "Don Candy": [("Papa", 220, "g"), ("Pollo crudo", 40, "g"), ("Chorizo", 40, "g"),
                  ("Salchicha", 1, "unidad"), ("Tocino", 30, "g"), ("Cheddar laminado", 40, "g"),
                  ("Carne molida", 45, "g"), ("AJUSTE: extras sin costear", 0.38, "$")],
    "Burger Clásica": [("Pan de hamburguesa", 1, "unidad"), ("Carne molida", 100, "g"),
                       ("Cheddar laminado", 20, "g"), ("AJUSTE: extras sin costear", 0.43, "$")],
    "Burger Sweet Onion": [("Pan de hamburguesa", 1, "unidad"), ("Carne molida", 100, "g"),
                           ("Cheddar laminado", 20, "g"), ("Cebolla caramelizada", 1, "porción"),
                           ("Aro de cebolla congelado", 25, "g"), ("AJUSTE: extras sin costear", 0.43, "$")],
    "Burger Completa": [("Pan de hamburguesa", 1, "unidad"), ("Carne molida", 100, "g"),
                        ("Cheddar laminado", 20, "g"), ("Tocino", 25, "g"), ("Huevo", 1, "unidad"),
                        ("AJUSTE: extras sin costear", 0.49, "$")],
    "CandyBurger": [("Pan de hamburguesa", 1, "unidad"), ("Carne molida", 100, "g"),
                    ("Cheddar laminado", 40, "g"), ("Tocino", 25, "g"), ("Huevo", 1, "unidad"),
                    ("Cebolla caramelizada", 1, "porción"), ("Aro de cebolla congelado", 25, "g"),
                    ("AJUSTE: extras sin costear", 0.43, "$")],
    "3 Tenders": [("Pollo crudo", 135, "g"), ("AJUSTE: extras sin costear", 0.60, "$")],
    "5 Tenders": [("Pollo crudo", 225, "g"), ("AJUSTE: extras sin costear", 0.71, "$")],
    "Adicional: Porción de papas": [("Papa", 110, "g"), ("AJUSTE: extras sin costear", 0.18, "$")],
    "Adicional: Cola + porción de papas": [("Papa", 110, "g"), ("Cola", 1, "unidad"),
                                           ("AJUSTE: extras sin costear", 0.18, "$")],
    "Adicional: Doble carne": [("Carne molida", 100, "g")],
    "Granizado de Máquina (sin alcohol)": [("Jarabe concentrado", 62.5, "ml"),
                                           ("Vaso 12oz + tapa + pitillo", 1, "set")],
    "Granizado de Máquina (con alcohol)": [("Jarabe concentrado", 62.5, "ml"),
                                           ("Vaso 12oz + tapa + pitillo", 1, "set"),
                                           ("Alcohol (aguardiente/vodka)", 30, "ml")],
    "Especial Oreo (sin alcohol)": [("Galleta Oreo", 4, "unidad"), ("Leche entera", 100, "ml"),
                                    ("Hielo", 150, "g"), ("Crema de leche", 20, "ml"),
                                    ("Vaso 12oz + tapa + pitillo", 1, "set")],
    "Especial Oreo (con alcohol)": [("Galleta Oreo", 4, "unidad"), ("Leche entera", 100, "ml"),
                                    ("Hielo", 150, "g"), ("Crema de leche", 20, "ml"),
                                    ("Vaso 12oz + tapa + pitillo", 1, "set"),
                                    ("Alcohol (aguardiente/vodka)", 30, "ml")],
    "Especial Nutella (sin alcohol)": [("Nutella", 30, "g"), ("Leche entera", 120, "ml"),
                                       ("Hielo", 150, "g"), ("Vaso 12oz + tapa + pitillo", 1, "set")],
    "Especial Nutella (con alcohol)": [("Nutella", 30, "g"), ("Leche entera", 120, "ml"),
                                       ("Hielo", 150, "g"), ("Vaso 12oz + tapa + pitillo", 1, "set"),
                                       ("Alcohol (aguardiente/vodka)", 30, "ml")],
    "Especial Café (sin alcohol)": [("Café soluble", 4, "g"), ("Leche condensada", 20, "ml"),
                                    ("Leche entera", 80, "ml"), ("Hielo", 160, "g"),
                                    ("Vaso 12oz + tapa + pitillo", 1, "set")],
    "Especial Café (con alcohol)": [("Café soluble", 4, "g"), ("Leche condensada", 20, "ml"),
                                    ("Leche entera", 80, "ml"), ("Hielo", 160, "g"),
                                    ("Vaso 12oz + tapa + pitillo", 1, "set"),
                                    ("Alcohol (aguardiente/vodka)", 30, "ml")],
}

r = 5
alt = False
for prod, items in R.items():
    fill = PatternFill("solid", fgColor="FAFAFA") if alt else None
    for i, (insumo, cant, unidad) in enumerate(items):
        rec.cell(row=r, column=1, value=prod if i == 0 else "")
        if i == 0:
            rec.cell(row=r, column=1).font = Font(bold=True, size=11)
        rec.cell(row=r, column=2, value=insumo)
        rec.cell(row=r, column=3, value=cant)
        rec.cell(row=r, column=4, value=unidad)
        f = rec.cell(row=r, column=5,
                     value=f"=IFERROR(C{r}*VLOOKUP(B{r},INSUMOS!$A${5}:$E${INS_LAST},5,FALSE),0)")
        f.number_format = MONEY2
        for col in range(1, 6):
            cell = rec.cell(row=r, column=col)
            cell.border = BORDER
            if fill:
                cell.fill = fill
        r += 1
    alt = not alt

REC_LAST = r - 1
rec.freeze_panes = "A5"

dv = DataValidation(type="list", formula1=f"=INSUMOS!$A$5:$A${INS_LAST}", allow_blank=True)
rec.add_data_validation(dv)
dv.add(f"B5:B{REC_LAST + 40}")

# fórmulas listas para filas nuevas
for rr in range(REC_LAST + 1, REC_LAST + 22):
    f = rec.cell(row=rr, column=5,
                 value=f"=IF(B{rr}=\"\",\"\",IFERROR(C{rr}*VLOOKUP(B{rr},INSUMOS!$A${5}:$E${INS_LAST},5,FALSE),0))")
    f.number_format = MONEY2

# ───────────────────────────── MENU ─────────────────────────────
mnu = wb.create_sheet("MENÚ")
mnu["A1"] = "MENÚ — precio, costo y margen"
mnu["A1"].font = Font(bold=True, size=14, color=NARANJA)
mnu["A2"] = "Solo tocás la columna PVP. Todo lo demás se calcula solo. Techo sano de food cost: 35%."
mnu["A2"].font = Font(italic=True, size=10, color="666666")
header(mnu, 4, ["CATEGORÍA", "PRODUCTO", "PVP", "COSTO", "FOOD COST", "MARGEN $", "SEMÁFORO"],
       [24, 36, 12, 13, 13, 13, 16])

MENU = [
    ("GRANIZADOS DE MÁQUINA", "Granizado de Máquina (sin alcohol)", 1.50),
    ("GRANIZADOS DE MÁQUINA", "Granizado de Máquina (con alcohol)", 2.00),
    ("GRANIZADOS ESPECIALES", "Especial Oreo (sin alcohol)", 3.00),
    ("GRANIZADOS ESPECIALES", "Especial Oreo (con alcohol)", 3.50),
    ("GRANIZADOS ESPECIALES", "Especial Nutella (sin alcohol)", 3.00),
    ("GRANIZADOS ESPECIALES", "Especial Nutella (con alcohol)", 3.50),
    ("GRANIZADOS ESPECIALES", "Especial Café (sin alcohol)", 3.00),
    ("GRANIZADOS ESPECIALES", "Especial Café (con alcohol)", 3.50),
    ("BANDEJAS", "Bandeja Clásica", 3.50),
    ("BANDEJAS", "Bandeja Completa", 4.50),
    ("BANDEJAS", "Bandeja Suprema", 5.50),
    ("CONOS — SOLITARIAS", "Solitaria de Pollo", 2.75),
    ("CONOS — SOLITARIAS", "Solitaria de Chorizo", 2.75),
    ("CONOS — LAS QUE TE GUSTAN", "Pollo Bacon", 3.75),
    ("CONOS — LAS QUE TE GUSTAN", "Chori Bacon", 3.75),
    ("CONOS — LAS QUE TE GUSTAN", "Carne Bacon", 3.75),
    ("CONOS — LA FAVORITA", "Don Candy", 5.50),
    ("HAMBURGUESAS", "Burger Clásica", 3.00),
    ("HAMBURGUESAS", "Burger Sweet Onion", 3.75),
    ("HAMBURGUESAS", "Burger Completa", 4.50),
    ("HAMBURGUESAS", "CandyBurger", 5.50),
    ("TENDERS", "3 Tenders", 3.50),
    ("TENDERS", "5 Tenders", 5.00),
    ("ADICIONALES", "Adicional: Porción de papas", 1.00),
    ("ADICIONALES", "Adicional: Cola + porción de papas", 1.50),
    ("ADICIONALES", "Adicional: Doble carne", 1.50),
    ("BEBIDAS", "", None),
]

r = 5
for cat, prod, pvp in MENU:
    mnu.cell(row=r, column=1, value=cat).font = Font(size=10, color="666666")
    mnu.cell(row=r, column=2, value=prod).font = Font(size=11, bold=True)
    c = mnu.cell(row=r, column=3, value=pvp)
    c.number_format = MONEY2
    c.fill = PatternFill("solid", fgColor="FFFDE7")
    mnu.cell(row=r, column=4,
             value=f"=SUMIF(RECETAS!$A${5}:$A${REC_LAST + 59},$B{r},RECETAS!$E${5}:$E${REC_LAST + 59})"
             ).number_format = MONEY2
    mnu.cell(row=r, column=5, value=f'=IF(OR($C{r}="",$C{r}=0),"",$D{r}/$C{r})').number_format = PCT
    mnu.cell(row=r, column=6, value=f'=IF($C{r}="","",$C{r}-$D{r})').number_format = MONEY2
    mnu.cell(row=r, column=7,
             value=f'=IF($E{r}="","",IF($E{r}<=0.3,"✅ SANO",IF($E{r}<=0.35,"🟡 LÍMITE","🔴 REVISAR")))')
    mnu.cell(row=r, column=7).alignment = Alignment(horizontal="center")
    for col in range(1, 8):
        mnu.cell(row=r, column=col).border = BORDER
    r += 1

MENU_LAST = r - 1

# filas extra listas
for rr in range(MENU_LAST + 1, MENU_LAST + 13):
    mnu.cell(row=rr, column=4,
             value=f'=IF($B{rr}="","",SUMIF(RECETAS!$A${5}:$A${REC_LAST + 59},$B{rr},RECETAS!$E${5}:$E${REC_LAST + 59}))'
             ).number_format = MONEY2
    mnu.cell(row=rr, column=5, value=f'=IF(OR($C{rr}="",$C{rr}=0),"",$D{rr}/$C{rr})').number_format = PCT
    mnu.cell(row=rr, column=6, value=f'=IF($C{rr}="","",$C{rr}-$D{rr})').number_format = MONEY2
    mnu.cell(row=rr, column=7,
             value=f'=IF($E{rr}="","",IF($E{rr}<=0.3,"✅ SANO",IF($E{rr}<=0.35,"🟡 LÍMITE","🔴 REVISAR")))')

LAST = MENU_LAST + 12
mnu.conditional_formatting.add(f"E5:E{LAST}",
    CellIsRule(operator="greaterThan", formula=["0.35"],
               fill=PatternFill("solid", fgColor="FFCDD2"), font=Font(bold=True, color="B71C1C")))
mnu.conditional_formatting.add(f"E5:E{LAST}",
    CellIsRule(operator="between", formula=["0.3", "0.35"],
               fill=PatternFill("solid", fgColor="FFF9C4"), font=Font(bold=True, color="F57F17")))
mnu.conditional_formatting.add(f"E5:E{LAST}",
    CellIsRule(operator="lessThan", formula=["0.3"],
               fill=PatternFill("solid", fgColor="C8E6C9"), font=Font(bold=True, color="1B5E20")))
mnu.freeze_panes = "A5"

OUT = "/private/tmp/claude-501/-Users-user-Projects-KEPLER/908a01ae-33c9-44ba-a2b6-ad9786af2284/scratchpad/CandyShots_Costos.xlsx"
wb.save(OUT)
print("OK ->", OUT)
print("INSUMOS ultima fila:", INS_LAST)
print("RECETAS ultima fila:", REC_LAST)
print("MENU ultima fila:", MENU_LAST)
